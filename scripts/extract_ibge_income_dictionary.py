#!/usr/bin/env python3
"""Extrai do dicionário oficial do IBGE as variáveis de renda usadas no projeto."""

import io
import json
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "raw" / "ibge" / "inspection" / "dicionario_renda_variaveis.json"
URL = (
    "https://ftp.ibge.gov.br/Censos/Censo_Demografico_2022/"
    "Agregados_por_Setores_Censitarios_Rendimento_do_Responsavel/"
    "dicionario_de_dados_renda_responsavel_20260508.xlsx"
)
VARIABLES = {"PERE0115_NOVA", "V06001", "V06002", "V06003", "V06004", "V06005", "V06006"}


def main() -> int:
    request = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=180) as response:
        raw = response.read()
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    matches = []
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = list(row)
            cells = [str(value).strip() for value in values if value is not None]
            found = sorted(
                variable for variable in VARIABLES
                if any(variable.upper() == cell.upper() for cell in cells)
            )
            if found:
                matches.append({
                    "sheet": worksheet.title,
                    "row": row_number,
                    "variables": found,
                    "values": values,
                })
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps({
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "source_url": URL,
        "variables_requested": sorted(VARIABLES),
        "matches": matches,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Definições localizadas: {len(matches)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
